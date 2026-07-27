#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BOOK = ROOT / "models/scenarios/S04/FINMODEL_VARSHAVKA_USALI_SCENARIO_S04_v3.0.0.xlsx"

wb = load_workbook(BOOK, data_only=False, read_only=True)
required = {"PRICE_REGISTER", "BREAKFAST_RECIPES", "BREAKFAST_COSTING"}
assert required.issubset(wb.sheetnames), required - set(wb.sheetnames)
assert len(wb.sheetnames) == 19, len(wb.sheetnames)

ws = wb["BREAKFAST_COSTING"]
for row in range(4, 20):
    assert ws[f"D{row}"].value == f"='BREAKFAST_RECIPES'!F{row}"
assert ws["G23"].value == "=E23+F23"
assert ws["G24"].value == "=E24+F24"
assert ws["G25"].value == "=E25+F25"
assert ws["G27"].value == "=E27+F27"
assert ws["I27"].value == "=G27/H27"
assert ws["B30"].value == "=G27*D30"

inputs = wb["01_ВВОД"]
assert inputs["D6"].value == "=DATE(YEAR(D5)+1,MONTH(D5),DAY(D5))-1"
assert inputs["F11"].value == "Требует подтверждения"
assert inputs["F15"].value == "Блокирующий input"
assert inputs["F24"].value == "Предварительный норматив"
assert inputs["F29"].value == "Предварительный расчёт"
assert inputs["F35"].value == "Предварительный норматив"
assert inputs["F42"].value == "Предварительный прокси"
assert inputs["F55"].value == "Предварительный норматив"
assert [inputs[f"F{row}"].value for row in (88, 89, 90)] == ["Требует основания"] * 3
assert [inputs[f"F{row}"].value for row in (141, 142, 143)] == [
    "Предварительно",
    "Предварительный расчёт",
    "Предварительный расчёт",
]
assert inputs["D141"].value == "=BREAKFAST_COSTING!$E$27"
assert inputs["D142"].value == "=BREAKFAST_COSTING!$G$27"
assert inputs["D143"].value == "=BREAKFAST_COSTING!$I$27"
assert inputs["G147"].value == "Кухня 131,282668 + напиток 60"
assert inputs["G148"].value == "Кухня 163,194668 + напиток 60"
assert inputs["G149"].value == "Кухня 133,483718 + напиток 60"

checks = wb["08_ПРОВЕРКИ"]
assert checks["A46"].value == "CHK.BREAKFAST.RESERVE"
assert "'03_ДОХОДЫ'!$P$22-7300" in checks["D46"].value
assert "'04_СЕБЕСТОИМОСТЬ'!$P$15" in checks["D46"].value
assert checks["A47"].value == "CHK.BREAKFAST.NORMS"
for row in range(4, 20):
    assert f"'BREAKFAST_COSTING'!D{row}-'BREAKFAST_RECIPES'!F{row}" in checks["D47"].value
recipes = wb["BREAKFAST_RECIPES"]
assert [recipes[f"J{row}"].value for row in range(11, 17)] == [45, 108, 86, 5, 5, 1]
assert sum(recipes[f"J{row}"].value for row in range(11, 17)) == 250
assert [recipes[f"F{row}"].value for row in range(11, 17)] == [45, 125, 100, 5, 5, 1]
assert checks["A48"].value == "CHK.BREAKFAST.OATMEAL_250"
assert checks["D48"].value == "=ABS(SUM('BREAKFAST_RECIPES'!J11:J16)-250)"
for row in range(49, 57):
    assert checks[f"A{row}"].value.startswith("CHK.")
assert checks["A49"].value == "CHK.CALENDAR.HORIZON"
assert checks["A50"].value == "CHK.BANQUET.COUNT"
assert checks["A51"].value == "CHK.BANQUET.MONTHLY"
assert checks["A52"].value == "CHK.BANQUET.MIX"
assert checks["A53"].value == "CHK.STAFF.CALENDAR"
assert checks["A54"].value == "CHK.COGS.PROXY_STATUS"
assert checks["A55"].value == "CHK.BREAKFAST.FC_ALIAS"
assert checks["A56"].value == "CHK.OPEN_INPUTS.VISIBLE"
wb_values = load_workbook(BOOK, data_only=True, read_only=True)
assert wb_values["08_ПРОВЕРКИ"]["F37"].value == "OK"

calendar = wb["02_КАЛЕНДАРЬ"]
assert calendar["A5"].value == "='01_ВВОД'!$D$5"
assert calendar["A6"].value == "=A5+1"
assert calendar["A369"].value == "=A368+1"
assert calendar["B5"].value.startswith("=12*(YEAR(A5)")
assert calendar["D5"].value == "=WEEKDAY(A5,2)"

assert inputs["D133"].value == "=SUM(D169:D175)"
assert inputs["D134"].value.startswith("=COUNTIF('02_КАЛЕНДАРЬ'!$D$5:$D$369,1)*D169")
assert [inputs[f"D{row}"].value for row in range(169, 176)] == [12, 12, 11, 11, 13, 14, 11]
assert inputs["D153"].value == "=65-(33-D152)"
assert inputs["D154"].value == "=656-(33-D152)*10"
assert inputs["D164"].value == "=(8*D160+8*D161+6*D162)/1000"

kitchen_program = wb["14_ПРОГРАММА_КУХНИ"]
assert kitchen_program["B9"].value == "='01_ВВОД'!$D$156*7"
assert kitchen_program["B10"].value == "='01_ВВОД'!$D$133"
assert kitchen_program["B11"].value == "='01_ВВОД'!$D$135"
assert kitchen_program["B12"].value == "='01_ВВОД'!$D$136"

pnl = wb["07_PNL_НАЛОГИ"]
assert pnl["A49"].value == "PNL.HOTEL.BREAKFAST.REV"
assert pnl["A52"].value == "PNL.HOTEL.DINNER.REV"
assert pnl["A55"].value == "PNL.HOTEL.TOTAL.REV"
assert pnl["P55"].value == "=SUM(D55:O55)"
assert pnl["P56"].value == "=SUM(D56:O56)"
assert pnl["P57"].value == "=SUM(D57:O57)"

# Independent benchmark from selected preliminary prices.
egg = 2 * (129.99 / 10) + 0.005 * (199.99 / 0.18) + 0.001 * 42.99
omelet = 0.126 * (361.50 / 0.9) + 0.050 * 146 + 0.005 * (199.99 / 0.18) + 0.001 * 42.99
oatmeal = 0.045 * (104.99 / 0.5) + 0.125 * 146 + 0.005 * 99.99 + 0.005 * (199.99 / 0.18) + 0.001 * 42.99
common = (2033.28 / 60) + 0.020 * (174.99 / 0.125) + 0.021 * 1800
weighted_full = 0.375 * (egg + common + 60) + 0.375 * (omelet + common + 60) + 0.25 * (oatmeal + common + 60)
assert abs(weighted_full - 203.79993027777778) < 1e-9
assert abs(weighted_full / 550 - 0.3705453277777778) < 1e-9
assert abs(weighted_full * 7300 - 1487739.4910277778) < 1e-6

formula_errors = []
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and any(token in value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                formula_errors.append(f"{sheet.title}!{cell.coordinate}:{value}")
assert not formula_errors, formula_errors[:20]

print({
    "workbook": str(BOOK),
    "sheets": len(wb.sheetnames),
    "weighted_full_cogs": weighted_full,
    "food_cost": weighted_full / 550,
    "annual_cogs": weighted_full * 7300,
    "formula_errors": 0,
})
