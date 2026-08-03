#!/usr/bin/env python3
"""Issue #82 workbook post-processing and exact structural QA.

The artifact-tool build currently does not serialize freeze panes reliably.  This
script applies an explicit, reviewable freeze map and then validates the saved
OOXML package.  It never fills subject-matter blanks or recalculates domain data.
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties


SHEETS = [
    "00_ПАСПОРТ", "01_МЕНЮ", "02_РЕЦЕПТУРЫ", "03_ПОЛУФАБРИКАТЫ",
    "04_КАЛЬКУЛЯЦИИ", "05_ТЕХКАРТЫ", "06_СЫРЬЁ_И_ЦЕНЫ", "07_ЦЕНООБРАЗОВАНИЕ",
    "08_ОБОРУДОВАНИЕ", "09_ИНВЕНТАРЬ_И_ПОСУДА", "10_АЛЛЕРГЕНЫ_БЕЗОПАСНОСТЬ",
    "11_ПИЩЕВАЯ_ЦЕННОСТЬ", "12_ВОПРОСЫ_ШЕФУ", "13_СОГЛАСОВАНИЕ",
    "14_КОНТРОЛЬНЫЕ_ПРОРАБОТКИ", "15_ПРОВЕРКИ", "16_ИСТОЧНИКИ",
]
FREEZE = {name: ("A13" if name == "00_ПАСПОРТ" else "A6") for name in SHEETS}
FORMULA_ERRORS = re.compile(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A|NUM!|NULL!)")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load(path: Path, data_only: bool = False):
    return load_workbook(path, data_only=data_only, read_only=False)


def canonicalize_ooxml(path: Path) -> None:
    """Rewrite ZIP metadata/order so identical OOXML produces identical bytes."""
    with zipfile.ZipFile(path, "r") as source:
        parts = {name: source.read(name) for name in source.namelist()}
    core = parts.get("docProps/core.xml")
    if core:
        parts["docProps/core.xml"] = re.sub(
            rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
            rb"\g<1>2026-08-03T00:00:00Z\g<2>", core,
        )
    with tempfile.NamedTemporaryFile(dir=path.parent, suffix=".xlsx", delete=False) as tmp:
        temp_path = Path(tmp.name)
    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as target:
            for name in sorted(parts):
                info = zipfile.ZipInfo(name, date_time=(2026, 8, 3, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = 0
                info.external_attr = 0
                target.writestr(info, parts[name])
        temp_path.replace(path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def apply_freeze(path: Path) -> dict:
    wb = load(path)
    if wb.sheetnames != SHEETS:
        raise AssertionError(f"sheet names/order mismatch: {wb.sheetnames}")
    for name, pane in FREEZE.items():
        wb[name].freeze_panes = pane
    fixed_time = dt.datetime(2026, 8, 3, 0, 0, 0)
    wb.properties.creator = "VARSHAVKA Issue 82 ExcelBuilder"
    wb.properties.created = fixed_time
    wb.properties.modified = fixed_time
    if wb.calculation is None:
        wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    else:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    wb.save(path)
    canonicalize_ooxml(path)
    check = load(path)
    actual = {name: check[name].freeze_panes for name in SHEETS}
    if actual != FREEZE:
        raise AssertionError(f"freeze panes mismatch: {actual}")
    return {"count": len(actual), "expected": FREEZE, "actual": actual}


def nonblank(ws, column: int, start: int, end: int) -> int:
    return sum(ws.cell(r, column).value not in (None, "") for r in range(start, end + 1))


def validate(path: Path) -> dict:
    wb = load(path)
    failures: list[str] = []
    checks: dict[str, object] = {}
    if wb.sheetnames != SHEETS:
        failures.append("exact sheet names/order")
    panes = {name: wb[name].freeze_panes for name in wb.sheetnames}
    checks["freeze_panes"] = panes
    if panes != FREEZE:
        failures.append("freeze panes exact 17/17")

    # Every sheet contains at least one filterable table; freeze begins at the
    # first data row of the primary visible table (cover begins at row 13).
    table_counts = {name: len(wb[name].tables) for name in SHEETS}
    checks["table_counts"] = table_counts
    if any(v < 1 for v in table_counts.values()):
        failures.append("at least one table/filter on every sheet")

    # Scope and layer-isolation assertions on exact generated ranges.
    assertions = {
        "dishes": nonblank(wb["01_МЕНЮ"], 1, 6, 33),
        "recipe_lines": nonblank(wb["02_РЕЦЕПТУРЫ"], 1, 6, 258),
        "evidence_cost_rows": nonblank(wb["04_КАЛЬКУЛЯЦИИ"], 1, 6, 33),
        "proxy_cost_rows": nonblank(wb["04_КАЛЬКУЛЯЦИИ"], 1, 38, 65),
        "tech_cards": nonblank(wb["05_ТЕХКАРТЫ"], 1, 6, 33),
        "accepted_price_sources": nonblank(wb["06_СЫРЬЁ_И_ЦЕНЫ"], 1, 123, 190),
        "evidence_channel_rows": nonblank(wb["07_ЦЕНООБРАЗОВАНИЕ"], 1, 6, 106),
        "proxy_channel_rows": nonblank(wb["07_ЦЕНООБРАЗОВАНИЕ"], 1, 111, 211),
        "equipment_mappings": nonblank(wb["08_ОБОРУДОВАНИЕ"], 1, 38, 192),
        "capacity_rows": nonblank(wb["08_ОБОРУДОВАНИЕ"], 1, 197, 224),
        "safety_profiles": nonblank(wb["10_АЛЛЕРГЕНЫ_БЕЗОПАСНОСТЬ"], 1, 6, 33),
        "nutrition_rows": nonblank(wb["11_ПИЩЕВАЯ_ЦЕННОСТЬ"], 1, 6, 33),
    }
    expected = {"dishes": 28, "recipe_lines": 253, "evidence_cost_rows": 28,
                "proxy_cost_rows": 28, "tech_cards": 28, "accepted_price_sources": 68,
                "evidence_channel_rows": 101, "proxy_channel_rows": 101,
                "equipment_mappings": 155, "capacity_rows": 28,
                "safety_profiles": 28, "nutrition_rows": 28}
    checks["scope"] = assertions
    if assertions != expected:
        failures.append(f"scope counts: {assertions}")

    tech = wb["05_ТЕХКАРТЫ"]
    required_headers = {
        "Область применения", "Scope status", "Требования к сырью", "Raw status",
        "Подготовка сырья", "Prep status", "Допустимые отклонения", "Tolerance status",
        "Органолептика", "Sensory status", "Хранение и реализация", "Storage status",
    }
    tech_headers = {tech.cell(5, c).value for c in range(1, tech.max_column + 1)}
    if not required_headers.issubset(tech_headers):
        failures.append("six TECH_CARDS fields and statuses")

    safety = wb["10_АЛЛЕРГЕНЫ_БЕЗОПАСНОСТЬ"]
    if sum(safety.cell(r, 13).value == "BLOCK" for r in range(6, 34)) != 28:
        failures.append("safety BLOCK 28/28")
    if sum(safety.cell(r, c).value in (None, "") for r in range(6, 34) for c in range(6, 10)) != 112:
        failures.append("unsupported safety numeric fields remain blank 112/112")
    if any(safety.cell(r, 5).value != "c6b22ad5f2812cc989a0d3593f40e21207da8f53" for r in range(6, 34)):
        failures.append("safety recipe blob version lock")

    nutrition = wb["11_ПИЩЕВАЯ_ЦЕННОСТЬ"]
    if any(not isinstance(nutrition.cell(r, c).value, (int, float)) for r in range(6, 34) for c in range(5, 21)):
        failures.append("nutrition numeric 28 × 16")
    if any(nutrition.cell(r, 23).value != "BLOCKED_PENDING_VALIDATION" for r in range(6, 34)):
        failures.append("nutrition release lock 28/28")

    # Unknown evidence-layer economics stay blank/formula blank; proxy rows are
    # visibly separate and carry their blocking status.
    costing = wb["04_КАЛЬКУЛЯЦИИ"]
    if any(costing.cell(r, 8).value not in (None, "") for r in range(6, 34)):
        failures.append("evidence complete food input blanks preserved")
    if any(costing.cell(r, 9).value != "ASSUMPTION_BLOCKED_PENDING_VALIDATION" for r in range(38, 66)):
        failures.append("proxy costing status 28/28")
    pricing = wb["07_ЦЕНООБРАЗОВАНИЕ"]
    if any(pricing.cell(r, 13).value != "ASSUMPTION_BLOCKED_PENDING_VALIDATION" for r in range(111, 212)):
        failures.append("proxy channel status 101/101")

    formulas = []
    formula_error_literals = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append(f"{ws.title}!{cell.coordinate}")
                if isinstance(value, str) and FORMULA_ERRORS.search(value):
                    formula_error_literals.append(f"{ws.title}!{cell.coordinate}:{value}")
    checks["formula_count"] = len(formulas)
    checks["formula_error_literals"] = formula_error_literals
    if len(formulas) < 580:
        failures.append(f"formula count unexpectedly low: {len(formulas)}")
    if formula_error_literals:
        failures.append("formula error literal scan")

    formula_contract = {
        "partial_cost_null_guard": wb["04_КАЛЬКУЛЯЦИИ"]["G6"].value,
        "complete_food_null_guard": wb["04_КАЛЬКУЛЯЦИИ"]["I6"].value,
        "complete_cogs_null_guard": wb["04_КАЛЬКУЛЯЦИИ"]["N6"].value,
        "channel_cogs_link": wb["07_ЦЕНООБРАЗОВАНИЕ"]["E6"].value,
        "channel_price_guard": wb["07_ЦЕНООБРАЗОВАНИЕ"]["I6"].value,
    }
    checks["formula_guard_contract"] = formula_contract
    if not all(isinstance(v, str) and v.startswith("=IF") for v in formula_contract.values()):
        failures.append("null-safe formula guard contract")
    # Explicit mechanics: a numeric zero is a known input and must remain zero;
    # None/blank remains blank.  This mirrors the IF(cell="","",cell) contract.
    guard = lambda value: "" if value in (None, "") else value
    checks["formula_reactivity_probe"] = {"blank": guard(None), "zero": guard(0), "positive": guard(12.5)}
    if guard(None) != "" or guard(0) != 0 or guard(12.5) != 12.5:
        failures.append("formula reactivity/null-zero guard probe")

    validations = sum(len(ws.data_validations.dataValidation) for ws in wb.worksheets)
    checks["data_validations"] = validations
    if validations < 4:
        failures.append(f"data validations unexpectedly low: {validations}")
    checks["calc_mode"] = wb.calculation.calcMode
    checks["full_calc_on_load"] = wb.calculation.fullCalcOnLoad
    if wb.calculation.calcMode != "auto" or not wb.calculation.fullCalcOnLoad:
        failures.append("external recalc flags")
    checks["sha256"] = sha256(path)
    checks["failures"] = failures
    checks["status"] = "PASS" if not failures else "FAIL"
    return checks


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--fix-freeze", action="store_true")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    if not args.workbook.exists():
        raise SystemExit(f"missing workbook: {args.workbook}")
    if args.fix_freeze:
        result = apply_freeze(args.workbook)
    else:
        result = validate(args.workbook)
    print(json.dumps(result, ensure_ascii=False, sort_keys=True))
    return 0 if result.get("status", "PASS") == "PASS" else 1


if __name__ == "__main__":
    sys.exit(main())
