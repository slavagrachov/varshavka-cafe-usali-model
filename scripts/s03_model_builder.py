from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SOURCE=Path('models/scenarios/S03/source/FINMODEL_VARSHAVKA_USALI_SCENARIO_S02_v0.2.0.xlsx')
TARGET=Path('models/scenarios/S03/FINMODEL_VARSHAVKA_USALI_SCENARIO_S03_v0.1.4.xlsx')
BRANCH='scenario/9-revised-dinner-pricing-and-opex'

def put(ws, values):
    for cell,value in values.items(): ws[cell]=value

def build_model():
    wb=load_workbook(SOURCE,data_only=False)
    wb.calculation.calcMode='auto'; wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True
    wb['00_РЕЗЮМЕ']['A1']='VARSHAVKA — Сценарий S03'
    i=wb['01_ВВОД']
    put(i,{
      'A117':'СЦЕНАРИЙ S03 — ПЕРЕСМОТР УЖИНОВ И OPEX',
      'A118':'S03_BANQ_ALA_SHARE','B118':'Сценарий S03','A119':'S03_DAY_END','B119':'Сценарий S03',
      'A120':'S03_DINNER_RATIO','A121':'S03_DINNER_CHECK','A122':'S03_DINNER_COGS',
      'C122':'Food cost ужина — как a-la carte','D122':'=D24','F122':'Расчёт',
      'G122':'Автоматически связано с ALA_COGS: food cost ужина равен food cost a-la carte.',
      'A123':'S03_DINNER_PACK','A124':'S03_DINNER_CASHLESS',
      'G118':'Наследовано из S02 / Issue #4; включено в S03 / Issue #9',
      'G119':'Наследовано из S02 / Issue #4; включено в S03 / Issue #9',
      'G120':'Наследовано из S02 / Issue #4; включено в S03 / Issue #9',
      'G121':'Средний чек ужина 1 000 руб.; требуется бизнес-подтверждение.',
      'G123':'Наследовано из S02; требует подтверждения.','G124':'Наследовано из S02; требует подтверждения.',
      'D27':0.25,'F27':'Подтверждено','G27':'Пользователь, 17.07.2026; скидка 25% от среднего чека a-la carte',
      'D28':'=D18*(1-D27)','F28':'Расчёт','G28':'Средний чек a-la carte × (1 − скидка) = 1 000 × 75% = 750 руб.',
      'G29':'30% / (1 − 25%)'})
    i['D122'].font=copy(i['D29'].font); i['D122'].fill=copy(i['D29'].fill)
    wb['03_ДОХОДЫ']['A2']='Сценарий S03 наследует бизнес-ланч S02. Чек ужина — 1 000 руб.; food cost ужина равен food cost a-la carte; пересмотрены OPEX-inputs.'
    wb['04_СЕБЕСТОИМОСТЬ']['C42']='Food cost ужинов 30% — как a-la carte'
    wb['07_PNL_НАЛОГИ']['C22']='Себестоимость ужинов — 30% как a-la carte'
    c=wb['08_ПРОВЕРКИ']
    for r in range(19,26):
        if isinstance(c.cell(r,1).value,str): c.cell(r,1).value=c.cell(r,1).value.replace('S02','S03')
    c['D25']='S03'; c['G25']='Отдельное имя файла сценария S03'
    for col in range(1,8):
        c.cell(26,col)._style=copy(c.cell(25,col)._style); c.cell(26,col).alignment=copy(c.cell(25,col).alignment)
    put(c,{'A26':'CHK.S03.DINNER.COGS','B26':'Food cost ужина связан с food cost a-la carte','C26':0,
           'D26':"=ABS('01_ВВОД'!$D$122-'01_ВВОД'!$D$24)",'E26':'=D26-C26','F26':'=IF(ABS(E26)<0.000001,"OK","ERROR")',
           'G26':'S03_DINNER_COGS = ALA_COGS','A27':'ВНИМАНИЕ: чек ужина и OPEX требуют подтверждения. Food cost ужина связан с ALA_COGS и равен 30%.'})
    old=wb['09_СРАВНЕНИЕ_S02']; old.title='09_СРАВНЕНИЕ_S03'
    put(old,{'A1':'Сценарий S03 — сравнение с базовой моделью v0.1.0','A2':f'Issue #9 | Ветка {BRANCH}','C4':'Сценарий S03, руб.','A18':'Ключевые правила S03'})
    if '10_СРАВНЕНИЕ_S02_S03' in wb.sheetnames: del wb['10_СРАВНЕНИЕ_S02_S03']
    ws=wb.create_sheet('10_СРАВНЕНИЕ_S02_S03'); ws.merge_cells('A1:H1'); ws.merge_cells('A2:H2')
    ws['A1']='Сравнение сценариев S02 и S03'; ws['A2']='S03 повышает чек ужина, связывает food cost ужина с a-la carte и пересматривает OPEX.'
    fin=[('Доходы кафе',38497700,"='07_PNL_НАЛОГИ'!P6"),('Доходы банкетов',3600000,"='07_PNL_НАЛОГИ'!P25"),
         ('Совокупные доходы',42097700,"='07_PNL_НАЛОГИ'!P33"),('Себестоимость кафе',14006708.485714287,"='07_PNL_НАЛОГИ'!P7"),
         ('Маржинальный доход кафе',21864413.51428571,"='07_PNL_НАЛОГИ'!P10"),('OPEX кафе',4917000,"=SUM('07_PNL_НАЛОГИ'!P12,'07_PNL_НАЛОГИ'!P14:P17)"),
         ('GOP кафе',-6214655.451231526,"='07_PNL_НАЛОГИ'!P18"),('Прибыль проекта до налога',-5539455.451231525,"='07_PNL_НАЛОГИ'!P34"),
         ('Налог к уплате',420977,"='07_PNL_НАЛОГИ'!P37"),('Чистая прибыль проекта',-5960432.451231525,"='07_PNL_НАЛОГИ'!P42"),
         ('Доходы ужинов',2007500,"='07_PNL_НАЛОГИ'!P21"),('Себестоимость ужинов',1003750,"='07_PNL_НАЛОГИ'!P22")]
    ws.append([]); ws.append(['Показатель','S02, руб.','S03, руб.','Изменение, руб.','Изменение, %'])
    for r,(label,s02,formula) in enumerate(fin,5):
        ws.cell(r,1,label); ws.cell(r,2,s02); ws.cell(r,3,formula); ws.cell(r,4,f'=C{r}-B{r}'); ws.cell(r,5,f'=IF(B{r}=0,0,D{r}/ABS(B{r}))')
    changes=[('CARPETS','Ковры',12000,5000,'Допущение'),('CASH_RENT','Аренда касс',9000,0,'Подтверждено'),
      ('COFFEE_RENT','Аренда кофемашины',10000,0,'Подтверждено'),('EQUIP_MAINT','Обслуживание оборудования',10000,0,'Подтверждено'),
      ('GREASE','Жироуловитель',8000,0,'Допущение'),('LAUNDRY','Стирка',15000,0,'Допущение'),('LEGAL','Юридические услуги',10000,0,'Допущение'),
      ('S03_DINNER_CHECK','Средний чек ужина',550,1000,'Рабочее допущение'),('SES','СЭС',15000,2000,'Подтверждено'),
      ('STAFF_MEALS','Питание персонала',0,15000,'Требует уточнения'),('VENT','Вентиляция',6000,2000,'Подтверждено'),
      ('S03_DINNER_COGS','Food cost ужина',0.5,0.3,'Расчёт / связано')]
    for col,v in enumerate(['Код input','Параметр','S02','S03','Изменение','Статус','Комментарий'],1): ws.cell(19,col,v)
    for r,(code,label,s02,s03,status) in enumerate(changes,20):
        vals=[code,label,s02,s03,s03-s02,status,'Изменение зафиксировано в S03 v0.1.4.']
        for col,v in enumerate(vals,1): ws.cell(r,col,v)
    for col,v in enumerate(['Техническая проверка','S02','Source v0.2.0','Публикационный S03'],1): ws.cell(33,col,v)
    for cell,val in {'A34':'Количество формул','B34':6588,'C34':6588,'D34':6589,'A35':'Контрольные строки','B35':21,'C35':21,'D35':22,'A36':'Формульные ошибки','B36':'0','C36':'0','D36':'0'}.items(): ws[cell]=val
    dark,teal,light,white='1F4E78','0F6B78','D9EAF7','FFFFFF'; thin=Side(style='thin',color='B7C9D6'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws['A1'].font=Font(size=16,bold=True,color=white); ws['A1'].fill=PatternFill('solid',fgColor=dark); ws['A2'].fill=PatternFill('solid',fgColor=light); ws['A2'].alignment=Alignment(wrap_text=True)
    for r,color in ((4,dark),(19,teal),(33,dark)):
        for col in range(1,8):
            cell=ws.cell(r,col); cell.font=Font(bold=True,color=white); cell.fill=PatternFill('solid',fgColor=color); cell.alignment=Alignment(horizontal='center',wrap_text=True); cell.border=border
    for r in list(range(5,17))+list(range(20,32))+list(range(34,37)):
        for col in range(1,8): ws.cell(r,col).border=border
    for col,width in {'A':31,'B':32,'C':17,'D':18,'E':17,'F':20,'G':64}.items(): ws.column_dimensions[col].width=width
    ws.freeze_panes='A4'; ws.auto_filter.ref='A19:G31'
    TARGET.parent.mkdir(parents=True,exist_ok=True); wb.save(TARGET)
    core=comparison=0
    for sheet in wb.worksheets:
        count=sum(1 for row in sheet.iter_rows() for cell in row if isinstance(cell.value,str) and cell.value.startswith('='))
        if sheet.title=='10_СРАВНЕНИЕ_S02_S03': comparison=count
        else: core+=count
    if len(wb.sheetnames)!=11 or core!=6589 or comparison!=36: raise AssertionError((len(wb.sheetnames),core,comparison))
    return core,core+comparison
